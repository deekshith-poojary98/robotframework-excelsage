import time
from src.ExcelSage import *
from assertpy import assert_that
from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import Color
import openpyxl as xl
import pytest
import shutil
import psutil
from pandas import DataFrame
import pandas as pd

exl = ExcelSage()

EXCEL_FILE_PATH = r".\data\sample.xlsx"
CSV_FILE_PATH = r".\data\sample.csv"
INVALID_EXCEL_FILE_PATH = r"..\data\sample1.xlsx"
NEW_EXCEL_FILE_PATH = r".\data\new_excel.xlsx"
INVALID_SHEET_NAME = "invalid[]sheet"
INVALID_CELL_ADDRESS = "AAAA1"
INVALID_ROW_INDEX = 1012323486523
INVALID_COLUMN_INDEX = 163841


@pytest.fixture
def setup_teardown(scope='function', autouse=False):
    yield
    if exl.active_workbook:
        exl.close_workbook()


def copy_test_excel_file(destination_file = r".\data\sample.xlsx"):
    source_file = r".\data\sample_original.xlsx"
    shutil.copy(source_file, destination_file)
    return destination_file


def delete_the_test_excel_file(files, max_retries=3, wait_time=2):
    def find_process_locking_file(file_path):
        processes = []
        for proc in psutil.process_iter(['pid', 'name', 'open_files']):
            try:
                for open_file in proc.info['open_files'] or []:
                    if open_file.path == file_path:
                        processes.append(proc)
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                pass
        return processes

    def kill_process(proc):
        try:
            proc.terminate()
            proc.wait(timeout=3)
            print(f"Killed process {proc.info['pid']} - {proc.info['name']}")
        except psutil.NoSuchProcess:
            print(f"Process {proc.info['pid']} no longer exists.")
        except psutil.AccessDenied:
            print(f"Permission denied to terminate process {proc.info['pid']}.")
        except psutil.TimeoutExpired:
            print(f"Timed out trying to terminate process {proc.info['pid']}.")

    for file_path in files:
        if os.path.exists(file_path):
            for attempt in range(max_retries):
                try:
                    os.chmod(file_path, 0o777)
                    os.remove(file_path)
                    break
                except PermissionError:
                    print(f"PermissionError on {file_path}. Attempt {attempt + 1}/{max_retries}.")
                    processes = find_process_locking_file(file_path)
                    if processes:
                        for index, proc in enumerate(processes, start=1):
                            print(index, proc)
                            kill_process(proc)
                        time.sleep(wait_time)
                    else:
                        print(f"No process found locking the file {file_path}. Trying to delete again.")
                        time.sleep(wait_time)


@pytest.fixture(scope="session", autouse=True)
def setup():
    copy_test_excel_file()
    yield
    delete_the_test_excel_file(files = [EXCEL_FILE_PATH, NEW_EXCEL_FILE_PATH, CSV_FILE_PATH])


def test_open_workbook_success(setup_teardown):
    workbook = exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    assert_that(workbook).is_instance_of(Workbook)


def test_open_workbook_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        exl.open_workbook(workbook_name=INVALID_EXCEL_FILE_PATH)

    assert_that(str(exc_info.value)).is_equal_to(f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path.")


def test_create_workbook_success(setup_teardown):
    sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
    workbook = exl.create_workbook(workbook_name=NEW_EXCEL_FILE_PATH, sheet_data=sheet_data, overwrite_if_exists=True)
    assert_that(workbook).is_instance_of(Workbook)


def test_create_workbook_file_already_exists(setup_teardown):
    with pytest.raises(FileAlreadyExistsError) as exc_info:
        exl.create_workbook(workbook_name=NEW_EXCEL_FILE_PATH)

    assert_that(str(exc_info.value)).is_equal_to(f"Unable to create workbook. The file '{NEW_EXCEL_FILE_PATH}' already exists. Set 'overwrite_if_exists=True' to overwrite the existing file.")


def test_create_workbook_type_error(setup_teardown):
    with pytest.raises(TypeError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], "John"]
        exl.create_workbook(workbook_name=NEW_EXCEL_FILE_PATH, sheet_data=sheet_data, overwrite_if_exists=True)

    assert_that(str(exc_info.value)).is_equal_to("Invalid row at index 3 of type 'str'. Each row in 'sheet_data' must be a list.")


def test_get_sheets_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    sheets = exl.get_sheets()
    assert_that(sheets).is_length(3).contains("Sheet1", "Offset_table", "Invalid_header")


def test_get_sheets_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.get_sheets()

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_add_sheet_success(setup_teardown):
    sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.add_sheet(sheet_name="Sheet3", sheet_data=sheet_data, sheet_pos=2)

    workbook = xl.load_workbook(filename=EXCEL_FILE_PATH)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(4).contains("Sheet1", "Offset_table", "Sheet3", "Invalid_header")
    sheet_to_delete = workbook['Sheet3']
    workbook.remove(sheet_to_delete)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_add_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.add_sheet(sheet_name="Sheet4")

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_add_sheet_sheet_exists(setup_teardown):
    with pytest.raises(SheetAlreadyExistsError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.add_sheet(sheet_name="Sheet1", sheet_data=sheet_data, sheet_pos=2)

    assert_that(str(exc_info.value)).is_equal_to("Sheet 'Sheet1' already exists.")


def test_add_sheet_sheet__invalid_position(setup_teardown):
    with pytest.raises(InvalidSheetPositionError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.add_sheet(sheet_name="Sheet5", sheet_data=sheet_data, sheet_pos=5)

    assert_that(str(exc_info.value)).is_equal_to("Invalid sheet position: 5. Maximum allowed is 3.")


def test_add_sheet_type_error(setup_teardown):
    with pytest.raises(TypeError) as exc_info:
        sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], "John"]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.add_sheet(sheet_name="Sheet6", sheet_data=sheet_data, sheet_pos=1)

    assert_that(str(exc_info.value)).is_equal_to("Invalid row at index 3 of type 'str'. Each row in 'sheet_data' must be a list.")


def test_delete_sheet_success(setup_teardown):
    workbook = xl.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.create_sheet(title="Sheet_to_delete")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    deleted_sheet = exl.delete_sheet(sheet_name="Sheet_to_delete")
    assert_that(deleted_sheet).is_equal_to("Sheet_to_delete")

    workbook = xl.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook.sheetnames
    assert_that(sheet).is_length(3).contains("Sheet1", "Offset_table", "Invalid_header")
    workbook.close()


def test_delete_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.delete_sheet(sheet_name=INVALID_SHEET_NAME)

    assert_that(str(exc_info.value)).is_equal_to(f"Sheet '{INVALID_SHEET_NAME}' doesn't exists.")


def test_fetch_sheet_data_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    sheet_data = exl.fetch_sheet_data(sheet_name="Offset_table",output_format="list", starting_cell="D6", ignore_empty_columns=True, ignore_empty_rows=True)
    assert_that(isinstance(sheet_data, list)).is_true()

    sheet_data = exl.fetch_sheet_data(sheet_name="Offset_table",output_format="dict", starting_cell="A1", ignore_empty_columns=True, ignore_empty_rows=True)
    assert_that(isinstance(sheet_data, list) and all(isinstance(item, dict) for item in sheet_data)).is_true()

    sheet_data = exl.fetch_sheet_data(sheet_name="Offset_table",output_format="dataframe", starting_cell="D6", ignore_empty_columns=True, ignore_empty_rows=True)
    assert_that(isinstance(sheet_data, DataFrame)).is_true()


def test_fetch_sheet_data_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.fetch_sheet_data(sheet_name="Offset_table",output_format="list", starting_cell=INVALID_CELL_ADDRESS, ignore_empty_columns=True, ignore_empty_rows=True)

    assert_that(str(exc_info.value)).is_equal_to(f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists.")


def test_fetch_sheet_data_invalid_output_format(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.fetch_sheet_data(sheet_name="Offset_table",output_format="invalid", starting_cell="D6", ignore_empty_columns=True, ignore_empty_rows=True)

    assert_that(str(exc_info.value)).is_equal_to("Invalid output format. Use 'list', 'dict', or 'dataframe'.")


def test_rename_sheet_success(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.create_sheet(title="Sheet_to_rename")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.rename_sheet(old_name="Sheet_to_rename", new_name="Sheet_renamed")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheets = workbook.sheetnames
    assert_that(sheets).contains("Sheet_renamed")
    sheet_to_delete = workbook["Sheet_renamed"]
    workbook.remove(sheet_to_delete)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_rename_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.rename_sheet(old_name="Sheet1", new_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_rename_sheet_sheet_exists(setup_teardown):
    with pytest.raises(SheetAlreadyExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.rename_sheet(old_name="Sheet1", new_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to("Sheet 'Offset_table' already exists.")


def test_rename_sheet_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.rename_sheet(old_name=INVALID_SHEET_NAME, new_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to(f"Sheet '{INVALID_SHEET_NAME}' doesn't exists.")


def test_get_cell_value_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    cell_value = exl.get_cell_value(sheet_name="Sheet1",  cell_name="A1")
    assert_that(cell_value).is_equal_to("First Name")


def test_get_cell_value_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_cell_value(sheet_name="Sheet1",  cell_name=INVALID_CELL_ADDRESS)

    assert_that(str(exc_info.value)).is_equal_to(f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists.")


def test_close_workbook_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.close_workbook()

    assert_that(exl.active_workbook).is_none()
    assert_that(exl.active_workbook_name).is_none()
    assert_that(exl.active_sheet).is_none()


def test_close_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.close_workbook()

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_save_workbook_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.add_sheet(sheet_name="New_sheet")
    exl.save_workbook()

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook.sheetnames
    assert_that(sheet).is_length(4).contains("New_sheet")
    sheet_to_delete = workbook["New_sheet"]
    workbook.remove(sheet_to_delete)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

def test_save_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.save_workbook()

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_set_active_sheet_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    active_sheet = exl.set_active_sheet(sheet_name="Offset_table")

    assert_that(active_sheet).is_equal_to("Offset_table")
    assert_that(str(exl.active_sheet)).contains("Offset_table")


def test_set_active_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.set_active_sheet(sheet_name="Offset_table")

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_set_active_sheet_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.set_active_sheet(sheet_name=INVALID_SHEET_NAME)

    assert_that(str(exc_info.value)).is_equal_to(f"Sheet '{INVALID_SHEET_NAME}' doesn't exists.")


def test_write_to_cell_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.write_to_cell(cell_name="AAA1", cell_value="New_value", sheet_name="Sheet1")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    cell_value = sheet["AAA1"].value
    assert_that(cell_value).is_equal_to("New_value")
    sheet["AAA1"] = None
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_write_to_cell_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.write_to_cell(cell_name=INVALID_CELL_ADDRESS, cell_value="New_value", sheet_name="Sheet1")

    assert_that(str(exc_info.value)).is_equal_to(f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists.")


def test_get_column_count_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    column_count = exl.get_column_count(sheet_name="Offset_table", starting_cell="D6", ignore_empty_columns=True)
    assert_that(column_count).is_equal_to(7)


def test_get_column_count_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_count(sheet_name="Offset_table", starting_cell=INVALID_CELL_ADDRESS)

    assert_that(str(exc_info.value)).is_equal_to(f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists.")


def test_get_row_count_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    row_count = exl.get_row_count(sheet_name="Offset_table", include_header=True, starting_cell="D6", ignore_empty_rows=True)
    assert_that(row_count).is_equal_to(52)


def test_get_row_count_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_row_count(sheet_name="Offset_table", starting_cell=INVALID_CELL_ADDRESS)

    assert_that(str(exc_info.value)).is_equal_to(f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists.")


def test_append_row_success(setup_teardown):
    data = ['Marisa', 'Pia', 'Female', 33, 'France', '21/05/2015', None, 1946]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.append_row(sheet_name="Sheet1", row_data=data)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook['Sheet1']

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        if row_values == data:
            row_to_delete = row[0].row
            sheet.delete_rows(row_to_delete)
            break
    else:
        assert False, "Row not appended"

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_row_success(setup_teardown):
    data = ['Mark', 'Pia', 'Male', 53, 'France', '11/09/2014', None, 1946]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.insert_row(sheet_name="Sheet1", row_data=data, row_index=10)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook['Sheet1']

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        if row_values == data:
            row_to_delete = row[0].row
            assert_that(row_to_delete).is_equal_to(10)
            sheet.delete_rows(row_to_delete)
            break
    else:
        assert False, "Row not inserted at index 10"

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_row_invalid_row_index(setup_teardown):
    with pytest.raises(InvalidRowIndexError) as exc_info:
        data = ['Mark', 'Pia', 'Male', 53, 'France', '11/09/2014', None, 1946]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.insert_row(row_data=data, row_index=INVALID_ROW_INDEX)

    assert_that(str(exc_info.value)).is_equal_to(f"Row index {INVALID_ROW_INDEX} is invalid or out of bounds. The valid range is 1 to 1048576.")


def test_delete_row_success(setup_teardown):
    data = ['Dee', 'Pia', 'Male', 53, 'France', '11/09/2014', None, 1946]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook['Sheet1']
    sheet.insert_rows(2)
    for col_index, value in enumerate(data, start=1):
        sheet.cell(row=2, column=col_index, value=value)

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.delete_row(row_index=2)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook['Sheet1']

    for row in sheet.iter_rows():
        row_values = [cell.value for cell in row]
        if row[0].row > 2:
            break
        if row_values == data:
            assert False, "Row not deleted at index 2"

    workbook.close()


def test_delete_row_invalid_row_index(setup_teardown):
    with pytest.raises(InvalidRowIndexError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.delete_row(row_index=INVALID_ROW_INDEX)

    assert_that(str(exc_info.value)).is_equal_to(f"Row index {INVALID_ROW_INDEX} is invalid or out of bounds. The valid range is 1 to 1048576.")


def test_append_column_success(setup_teardown):
    data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.append_column(sheet_name="Sheet1", col_data=data)

    expected_values = ["data1", "data2", "data3", "data4", "data5"]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook['Sheet1']
    last_column = sheet.max_column
    header = sheet.cell(row=1, column=last_column).value
    assert_that(header).is_equal_to("New Column")
    new_column_values = [sheet.cell(row=row, column=last_column).value for row in range(2, 7)]
    assert_that(new_column_values).is_equal_to(expected_values)

    sheet.delete_cols(last_column)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_column_success(setup_teardown):
    data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.insert_column(sheet_name="Sheet1", col_data=data, col_index=1)

    expected_values = ["data1", "data2", "data3", "data4", "data5"]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook['Sheet1']
    header = sheet.cell(row=1, column=1).value

    assert_that(header).is_equal_to("New Column")
    new_column_values = [sheet.cell(row=row, column=1).value for row in range(2, 7)]
    assert_that(new_column_values).is_equal_to(expected_values)

    sheet.delete_cols(1)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_insert_column_invalid_column_index(setup_teardown):
    with pytest.raises(InvalidColumnIndexError) as exc_info:
        data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.insert_column(sheet_name="Sheet1", col_data=data, col_index=INVALID_COLUMN_INDEX)

    assert_that(str(exc_info.value)).is_equal_to(f"Column index {INVALID_COLUMN_INDEX} is invalid or out of bounds. The valid range is 1 to 16384.")


def test_delete_column_success(setup_teardown):
    data = ["New Column", "data1", "data2", "data3", "data4", "data5"]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    sheet.insert_cols(1)

    for row_index, value in enumerate(data, start=1):
        sheet.cell(row=row_index, column=1, value=value)

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.delete_column(sheet_name="Sheet1", col_index=1)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook['Sheet1']
    header = sheet.cell(row=1, column=1).value
    workbook.close()

    assert_that(header).is_not_equal_to("New Column")


def test_delete_column_invalid_column_index(setup_teardown):
    with pytest.raises(InvalidColumnIndexError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.delete_column(sheet_name="Sheet1", col_index=INVALID_COLUMN_INDEX)

    assert_that(str(exc_info.value)).is_equal_to(f"Column index {INVALID_COLUMN_INDEX} is invalid or out of bounds. The valid range is 1 to 16384.")


@pytest.mark.parametrize("column_name, expected_length, expected_values, output_format", [
    ("A", 52, ["Tommie", "Nereida", "Stasia"], "list"),
    (["A", "B"], 52, [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]], "list"),
    ("First Name", 52, ["Tommie", "Nereida", "Stasia"], "list"),
    (["First Name", "Last Name"], 52, [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]], "list"),

    ("A", 52, [["Tommie", "Nereida", "Stasia"]], "dict"),
    (["A", "B"], 52, [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]], "dict"),
    ("First Name", 52, [["Tommie", "Nereida", "Stasia"]], "dict"),
    (["First Name", "Last Name"], 52, [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]], "dict"),

    ("A", 52, ["Tommie", "Nereida", "Stasia"], "DataFrame"),
    (["A", "B"], 52, [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]], "DataFrame"),
    ("First Name", 52, ["Tommie", "Nereida", "Stasia"], "DataFrame"),
    (["First Name", "Last Name"], 52, [["Tommie", "Nereida", "Stasia"], ["Mccrystal", "Partain", "Hanner"]], "DataFrame")
])
def test_get_column_values_success(setup_teardown, column_name, expected_length, expected_values, output_format):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    column_values = exl.get_column_values(column_names_or_letters=column_name, sheet_name="Offset_table", starting_cell="D6", output_format=output_format)

    if isinstance(column_name, str) or output_format in ["dict", "DataFrame"]:
        assert_that(isinstance(column_values, eval(output_format))).is_true()

        if output_format == "list":
            assert_that(column_values).is_length(expected_length).contains(*expected_values)
        elif output_format == "dict":
            keys = list(column_values.keys())
            for i, key in enumerate(keys):
                values = column_values.get(key)
                assert_that(values).is_length(expected_length).contains(*expected_values[i])
        else:
            row_count, column_count = column_values.shape
            assert_that(row_count).is_equal_to(expected_length)
            column_headers = column_values.columns.tolist()

            if column_count == 1:
                assert_that(column_values[column_headers[0]].tolist()).contains(*expected_values)
            else:
                assert_that(column_values[column_headers[0]].tolist()).contains(*expected_values[0])
                assert_that(column_values[column_headers[1]].tolist()).contains(*expected_values[1])
    else:
        assert_that(isinstance(column_values, eval(output_format))).is_true()
        assert_that(column_values).is_length(2)
        for sublist in column_values:
            assert_that(isinstance(sublist, list)).is_true()
            assert_that(sublist).is_length(expected_length)


def test_get_column_values_invalid_output_format(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(column_names_or_letters="A", sheet_name="Offset_table", starting_cell="D6", output_format="invalid")

    assert_that(str(exc_info.value)).is_equal_to("Invalid output format. Use 'list', 'dict', or 'dataframe'.")


def test_get_column_values_invalid_header(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(column_names_or_letters="G", sheet_name="Invalid_header", starting_cell="A1", output_format="list")

    assert_that(str(exc_info.value)).is_equal_to("Column letter 'G' does not have a valid string header: '1234' found.")


@pytest.mark.parametrize("invalid_columns",["AAAA", "Invalid Column Header"])
def test_get_column_values_invalid_column(setup_teardown, invalid_columns):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(column_names_or_letters=invalid_columns, sheet_name="Sheet1", starting_cell="A1", output_format="list")

    assert_that(str(exc_info.value)).is_equal_to(f"Invalid column name or letter: '{invalid_columns}'")


def test_get_column_values_column_out_of_bound(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_values(column_names_or_letters="Z", sheet_name="Sheet1", starting_cell="A55", output_format="list")

    assert_that(str(exc_info.value)).is_equal_to(f"Column letter 'Z' is out of bounds for the provided sheet.")


@pytest.mark.parametrize("row_index, expected_values, expected_length, output_format", [
    (2,["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574], 8, "list"),
    ([2, 3],[["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574], ["Francesca", "Beaudreau", "Female", 21, "France", "15/10/2017", 5412]], 8, "list"),
    (2,[["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574]], 8, "dict"),
    ([2, 3],[["Lester", "Prothro", "Male", 20, "France", "15/10/2017", None, 6574], ["Francesca", "Beaudreau", "Female", 21, "France", "15/10/2017", 5412]], 8, "dict")
 ])
def test_get_row_values_success(setup_teardown, row_index, expected_length, expected_values, output_format):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    row_value = exl.get_row_values(sheet_name="Sheet1", row_indices=row_index, output_format=output_format)
    assert_that(isinstance(row_value, eval(output_format))).is_true()

    if isinstance(row_index, int) or output_format == "dict":
        if output_format == "list":
            assert_that(row_value).is_length(expected_length).contains(*expected_values)
        else:
            keys = list(row_value.keys())
            for i, key in enumerate(keys):
                values = row_value.get(key)
                assert_that(values).is_length(expected_length).contains(*expected_values[i])
    else:
        assert_that(isinstance(row_value, eval(output_format))).is_true()
        assert_that(row_value).is_length(2)
        for sublist in row_value:
            assert_that(isinstance(sublist, list)).is_true()
            assert_that(sublist).is_length(expected_length)


def test_get_row_values_invalid_output_format(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_row_values(sheet_name="Sheet1", row_indices=2, output_format="invalid")

    assert_that(str(exc_info.value)).is_equal_to("Invalid output format. Use 'list' or 'dict'.")


def test_get_row_values_invalid_row_index(setup_teardown):
    with pytest.raises(InvalidRowIndexError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_row_values(sheet_name="Sheet1", row_indices=INVALID_ROW_INDEX, output_format="list")

    assert_that(str(exc_info.value)).is_equal_to(f"Row index {INVALID_ROW_INDEX} is invalid or out of bounds. The valid range is 1 to 1048576.")


def test_protect_sheet_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.protect_sheet(sheet_name="Invalid_header", password="password")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    assert_that(sheet.protection.sheet).is_true()
    sheet.protection.set_password("password")
    sheet.protection.sheet = False
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_protect_sheet_sheet_already_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    sheet.protection.set_password("password")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(SheetAlreadyProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.protect_sheet(sheet_name="Invalid_header", password="password")

    assert_that(str(exc_info.value)).is_equal_to("The sheet 'Invalid_header' is already protected and cannot be protected be again.")


def test_unprotect_sheet_success(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    sheet.protection.set_password("password")
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.unprotect_sheet(sheet_name="Invalid_header", password="password")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    assert_that(sheet.protection.sheet).is_false()
    workbook.close()


def test_unprotect_sheet_sheet_not_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Invalid_header"]
    sheet.protection.set_password("password")
    sheet.protection.sheet = False
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(SheetNotProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.unprotect_sheet(sheet_name="Invalid_header", password="password")

    assert_that(str(exc_info.value)).is_equal_to("The sheet 'Invalid_header' is not currently protected and cannot be unprotected.")


def test_protect_workbook_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.protect_workbook(password="password", protect_sheets=True)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    assert_that(workbook.security.lockStructure).is_true()
    workbook.security.lockStructure = False

    for sheet in workbook.worksheets:
        if sheet.protection.sheet:
            sheet.protection.sheet = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_protect_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.protect_workbook(password="password", protect_sheets=True)

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_protect_workbook_workbook_already_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    protection = WorkbookProtection()
    protection.workbookPassword = "password"
    protection.lockStructure = True
    protection.lockWindows = True

    workbook.security = protection

    for sheet in workbook.worksheets:
        sheet.protection = SheetProtection(password="password")
        sheet.protection.sheet = True
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.objects = False
        sheet.protection.scenarios = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(WorkbookAlreadyProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.protect_workbook(protect_sheets=True, password="password")

    assert_that(str(exc_info.value)).is_equal_to("The workbook is already protected and cannot be protected be again.")


def test_unprotect_workbook_success(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    protection = WorkbookProtection()
    protection.workbookPassword = "password"
    protection.lockStructure = True
    protection.lockWindows = True

    workbook.security = protection

    for sheet in workbook.worksheets:
        sheet.protection = SheetProtection(password="password")
        sheet.protection.sheet = True
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.objects = False
        sheet.protection.scenarios = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.unprotect_workbook(unprotect_sheets=True)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    assert_that(workbook.security and workbook.security.lockStructure).is_false()
    workbook.close()


def test_unprotect_workbook_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.unprotect_workbook(unprotect_sheets=True)

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_unprotect_workbook_workbook_not_protected(setup_teardown):
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.security.lockStructure = False

    for sheet in workbook.worksheets:
        if sheet.protection.sheet:
            sheet.protection.sheet = False

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    with pytest.raises(WorkbookNotProtectedError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.unprotect_workbook(unprotect_sheets=True)

    assert_that(str(exc_info.value)).is_equal_to("The workbook is not currently protected and cannot be unprotected.")


def test_clear_sheet_success(setup_teardown):
    sheet_data = [["Name", "Age"], ["Dee", 26], ["Mark", 56], ["John", 30]]
    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    workbook.create_sheet(title="Clear_sheet", index=1)
    sheet = workbook["Clear_sheet"]

    for row in sheet_data:
        sheet.append(row)

    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.clear_sheet(sheet_name="Clear_sheet")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Clear_sheet"]

    is_empty = True
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=sheet.max_column, values_only=True):
        if any(cell is not None for cell in row):
            is_empty = False
            break

    assert_that(is_empty).is_true()
    workbook.remove(sheet)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_copy_sheet_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    exl.copy_sheet(source_sheet_name="Sheet1", new_sheet_name="Copied_sheet")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(4).contains("Sheet1", "Offset_table", "Invalid_header", "Copied_sheet")

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet1 = workbook["Sheet1"]
    sheet2 = workbook["Copied_sheet"]

    if sheet1.max_row != sheet2.max_row or sheet1.max_column != sheet2.max_column:
        assert False, "Rows/Columns not matching"

    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column + 1):
            cell1 = sheet1.cell(row=row, column=col).value
            cell2 = sheet2.cell(row=row, column=col).value

            if cell1 != cell2:
                assert False, "Cells not matching"

    workbook.remove(sheet2)
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()


def test_copy_sheet_workbook_not_open(setup_teardown):
    with pytest.raises(WorkbookNotOpenError) as exc_info:
        exl.copy_sheet(source_sheet_name="Sheet1", new_sheet_name="Copied_sheet")

    assert_that(str(exc_info.value)).is_equal_to("Workbook isn't open. Please open the workbook first.")


def test_copy_sheet_invalid_sheet_name(setup_teardown):
    with pytest.raises(InvalidSheetNameError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.copy_sheet(source_sheet_name="Sheet1", new_sheet_name=INVALID_SHEET_NAME)

    assert_that(str(exc_info.value)).is_equal_to(f"The sheet name '{INVALID_SHEET_NAME}' is invalid.")


def test_copy_sheet_doesnt_exists(setup_teardown):
    with pytest.raises(SheetDoesntExistsError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.copy_sheet(source_sheet_name=INVALID_SHEET_NAME, new_sheet_name="Copied_sheet")

    assert_that(str(exc_info.value)).is_equal_to(f"Sheet '{INVALID_SHEET_NAME}' doesn't exists.")


@pytest.mark.parametrize("value, occurence, expected_value",
                         [("Male", "first", ("str", "C2")),
                          ("Lester", "all", ("list", ["A2", "A10"])),
                          ("Invalid_value", "first", (None, None))
                          ])
def test_find_value_success(setup_teardown, value, occurence, expected_value):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    cell = exl.find_value(sheet_name="Sheet1", value=value, occurence=occurence)

    if expected_value[0] is None:
        assert_that(cell).is_none()
    else:
        assert_that(isinstance(cell, eval(expected_value[0]))).is_true()
        assert_that(cell).is_equal_to(expected_value[1])


def test_find_value_invalid_occurence(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_value(sheet_name="Sheet1", value="value", occurence="invalid_occurence")

    assert_that(str(exc_info.value)).is_equal_to("Invalid occurence, use either 'first' or 'all'.")


@pytest.mark.parametrize("occurence, expected_value", [("first", ("str", "A14")),
                                                       ("all", ("list", ["A18", "A19"])),
                                                       ("first", (None, None))])
def test_find_and_replace_success(setup_teardown, occurence, expected_value):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    cell = exl.find_and_replace(sheet_name="Sheet1", old_value="Marcel", new_value="Mark", occurence=occurence)

    if expected_value[0] is None:
        assert_that(cell).is_none()
    else:
        assert_that(isinstance(cell, eval(expected_value[0]))).is_true()
        assert_that(cell).is_equal_to(expected_value[1])

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook.active

    if isinstance(cell, list):
        for i in cell:
            assert_that(sheet[i].value).is_equal_to("Mark")
    elif isinstance(cell, str):
        assert_that(sheet[cell].value).is_equal_to("Mark")

    workbook.close()


def test_find_and_replace_value_invalid_occurence(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.find_and_replace(sheet_name="Sheet1", old_value="Marcel", new_value="Mark", occurence="invalid_occurence")

    assert_that(str(exc_info.value)).is_equal_to("Invalid occurence, use either 'first' or 'all'.")


def test_format_cell_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    alignment_config = {
    "vertical": "center",
    "horizontal": "left"
    }

    border_config = {
        "left": True,
        "right": True,
        "top": True,
        "bottom": True,
        "style": "thin",
        "color": "#FF0000"
    }
    exl.format_cell(
        sheet_name="Sheet1", cell_name="C3", font_size=12, font_color="#FF0000",
        alignment=alignment_config, wrap_text=True, bg_color="#FFFF00", cell_width=120, cell_height=25,
        font_name="Arial", bold=True, italic=True, strike_through=True, underline=True, border=border_config, auto_fit_height=False, auto_fit_width=False
    )

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Sheet1"]
    cell = sheet["C3"]

    cell_properties = {
        "cell_name": cell.coordinate,
        "font_size": cell.font.size,
        "font_color": cell.font.color.rgb if cell.font.color else None,
        "alignment_horizontal": cell.alignment.horizontal,
        "alignment_vertical": cell.alignment.vertical,
        "wrap_text": cell.alignment.wrap_text,
        "bg_color": cell.fill.start_color.rgb if cell.fill.start_color else None,
        "cell_width": sheet.column_dimensions[cell.column_letter].width,
        "cell_height": sheet.row_dimensions[cell.row].height,
        "font_name": cell.font.name,
        "bold": cell.font.bold,
        "italic": cell.font.italic,
        "underline": cell.font.underline,
        "strike_through": cell.font.strike,
        "border": {
            "top": cell.border.top.style if cell.border.top else None,
            "bottom": cell.border.bottom.style if cell.border.bottom else None,
            "left": cell.border.left.style if cell.border.left else None,
            "right": cell.border.right.style if cell.border.right else None,
        }
    }
    assert_that(cell_properties['font_size']).is_equal_to(12.0)
    assert_that(cell_properties['font_color']).is_equal_to("FFFF0000")
    assert_that(cell_properties["alignment_horizontal"]).is_equal_to("left")
    assert_that(cell_properties["alignment_vertical"]).is_equal_to("center")
    assert_that(cell_properties["wrap_text"]).is_true()
    assert_that(cell_properties["bg_color"]).is_equal_to("FFFFFF00")
    assert_that(cell_properties["cell_width"]).is_equal_to(120.0)
    assert_that(cell_properties["cell_height"]).is_equal_to(25.0)
    assert_that(cell_properties["font_name"]).is_equal_to("Arial")
    assert_that(cell_properties["bold"]).is_true()
    assert_that(cell_properties["italic"]).is_true()
    assert_that(cell_properties["underline"]).is_equal_to("single")
    assert_that(cell_properties["strike_through"]).is_true()
    assert_that(cell_properties["border"]).is_equal_to({'top': 'thin', 'bottom': 'thin', 'left': 'thin', 'right': 'thin'})
    workbook.close()

    exl.format_cell(sheet_name="Offset_table", cell_name="D6", auto_fit_height=True, auto_fit_width=True)

    workbook = excel.load_workbook(filename=EXCEL_FILE_PATH)
    sheet = workbook["Offset_table"]
    cell = sheet["D6"]
    cell_value = str(cell.value) if cell.value else ""
    col_letter = get_column_letter(cell.column)
    max_length = max(len(cell_value), len(col_letter))
    computed_width = max_length + 2
    max_line_count = cell_value.count('\n') + 1
    computed_height = max(15, max_line_count * 15)
    cell_properties = {
        "auto_width": computed_width,
        "auto_height": computed_height,
    }

    assert_that(cell_properties["auto_width"]).is_equal_to(12)
    assert_that(cell_properties["auto_height"]).is_equal_to(15)
    workbook.close()


def test_format_cell_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name=INVALID_CELL_ADDRESS, font_size=10)

    assert_that(str(exc_info.value)).is_equal_to(f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists.")


def test_format_cell_invalid_font_color(setup_teardown):
    with pytest.raises(InvalidColorError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", font_color="invalid_color")

    assert_that(str(exc_info.value)).is_equal_to("Invalid font color: 'invalid_color'. Use valid hex color in #RRGGBB format.")


def test_format_cell_invalid_bg_color(setup_teardown):
    with pytest.raises(InvalidColorError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", bg_color="invalid_color")

    assert_that(str(exc_info.value)).is_equal_to("Invalid background color: 'invalid_color'. Use valid hex color in #RRGGBB format.")


def test_format_cell_invalid_border_color(setup_teardown):
    with pytest.raises(InvalidColorError) as exc_info:
        border_config = {
            "left": True,
            "right": True,
            "top": True,
            "bottom": True,
            "style": "thin",
            "color": "invalid_color"
        }
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", border=border_config)

    assert_that(str(exc_info.value)).is_equal_to("Invalid border color: 'invalid_color'. Use valid hex color in #RRGGBB format.")


def test_format_cell_invalid_horizontal_alignment(setup_teardown):
    with pytest.raises(InvalidAlignmentError) as exc_info:
        alignment_config = {
            "vertical": "center",
            "horizontal": "invalid_alignment"
            }
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", alignment=alignment_config)

    assert_that(str(exc_info.value)).is_equal_to("Invalid horizontal alignment: 'invalid_alignment'. Allowed values are ['left', 'center', 'right'].")


def test_format_cell_invalid_vertical_alignment(setup_teardown):
    with pytest.raises(InvalidAlignmentError) as exc_info:
        alignment_config = {
            "vertical": "invalid_alignment",
            "horizontal": "left"
            }
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", alignment=alignment_config)

    assert_that(str(exc_info.value)).is_equal_to("Invalid vertical alignment: 'invalid_alignment'. Allowed values are ['top', 'center', 'bottom'].")


def test_format_cell_invalid_border_style(setup_teardown):
    with pytest.raises(InvalidBorderStyleError) as exc_info:
        border_config = {
            "left": True,
            "right": True,
            "top": True,
            "bottom": True,
            "style": "invalid_style"
        }
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.format_cell(sheet_name="Offset_table", cell_name="D6", border=border_config)

    assert_that(str(exc_info.value)).is_equal_to("Invalid border style: 'invalid_style'. Allowed values are ['dashDot', 'dashDotDot', 'dashed', 'dotted', 'double', 'hair', 'medium', 'mediumDashDot', 'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'].")


def test_get_column_headers_success(setup_teardown):
    exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
    headers = exl.get_column_headers(sheet_name="Offset_table", starting_cell="D6")
    assert_that(headers).is_length(7).contains("First Name", "Last Name", "Gender",	"Country", "Age", "Date", "Salary")


def test_get_column_headers_invalid_cell_address(setup_teardown):
    with pytest.raises(InvalidCellAddressError) as exc_info:
        exl.open_workbook(workbook_name=EXCEL_FILE_PATH)
        exl.get_column_headers(sheet_name="Offset_table", starting_cell=INVALID_CELL_ADDRESS)

    assert_that(str(exc_info.value)).is_equal_to(f"Cell '{INVALID_CELL_ADDRESS}' doesn't exists.")


def test_export_to_csv_success(setup_teardown):
    output_filename = exl.export_to_csv(filename=EXCEL_FILE_PATH, output_filename=CSV_FILE_PATH, sheet_name="Sheet1", overwrite_if_exists=True)

    assert_that(output_filename).is_equal_to(CSV_FILE_PATH)
    assert_that(os.path.exists(CSV_FILE_PATH)).is_true()


def test_export_to_csv_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        exl.export_to_csv(filename=INVALID_EXCEL_FILE_PATH, output_filename=CSV_FILE_PATH, sheet_name="Sheet1", overwrite_if_exists=True)

    assert_that(str(exc_info.value)).is_equal_to(f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path.")


def test_export_to_csv_file_already_exists(setup_teardown):
    with pytest.raises(FileAlreadyExistsError) as exc_info:
        exl.export_to_csv(filename=EXCEL_FILE_PATH, output_filename=CSV_FILE_PATH, sheet_name="Sheet1", overwrite_if_exists=False)

    assert_that(str(exc_info.value)).is_equal_to(f"Unable to create workbook. The file '{CSV_FILE_PATH}' already exists. Set 'overwrite_if_exists=True' to overwrite the existing file.")


def test_merge_excels_multi_sheet_success(setup_teardown):
    NEW_FILE = copy_test_excel_file(destination_file = r".\data\sample2.xlsx")
    list_of_files = [EXCEL_FILE_PATH, NEW_FILE]
    output_file = r".\data\merged_file_multi_sheets.xlsx"
    exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="multiple_sheets", skip_bad_rows=True)
    assert_that(os.path.exists(output_file)).is_true()

    workbook = excel.load_workbook(filename=output_file)
    sheets = workbook.sheetnames
    workbook.close()
    assert_that(sheets).is_length(6).contains("Sheet1_sample", "Offset_table_sample2", "Offset_table_sample", "Invalid_header_sample", "Invalid_header_sample2", "Sheet1_sample2")
    delete_the_test_excel_file(files=[output_file, NEW_FILE])


def test_merge_excels_single_sheet_success(setup_teardown):
    data = {
        r".\data\single_sheet_workbook1.xlsx": [
            ["Name", "Age"],
            ["Mark", 25],
            ["John", 30]
        ],
        r".\data\single_sheet_workbook2.xlsx": [
            ["Name", "Age"],
            ["Dee", 26],
            ["Alex", 40]
        ]
    }

    def create_workbook(filename, data):
        workbook = Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(filename)
        workbook.close()

    for filename, workbook_data in data.items():
        create_workbook(filename, workbook_data)

    list_of_files = [r".\data\single_sheet_workbook1.xlsx", r".\data\single_sheet_workbook2.xlsx"]
    output_file = r".\data\merged_file_single_sheet.xlsx"
    exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="single_sheet", skip_bad_rows=True)
    assert_that(os.path.exists(output_file)).is_true()

    expected_data = [["Name", "Age"],["Mark", 25],["John", 30],["Dee", 26],["Alex", 40]]
    workbook = excel.load_workbook(filename=output_file)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(1)
    sheet = workbook[sheets[0]]
    assert_that(sheet.max_row).is_equal_to(5)
    assert_that(sheet.max_column).is_equal_to(2)

    for row_index, expected_row in enumerate(expected_data, start=1):
        for col_index, expected_value in enumerate(expected_row, start=1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            if cell_value != expected_value:
                assert False, f"Data mismatch ({cell_value} != {expected_value})"

    workbook.close()
    delete_the_test_excel_file(files=[r".\data\single_sheet_workbook1.xlsx", r".\data\single_sheet_workbook2.xlsx", output_file, ])


def test_merge_excels_sheet_wise_success(setup_teardown):
    data = {
        r".\data\sheet_wise_workbook1.xlsx": {
            "Sheet1": [
                ["Name", "Age"],
                ["Mark", 25],
                ["John", 30]
            ],
            "Sheet2": [
                ["City", "Country"],
                ["New York", "USA"],
                ["London", "UK"]
            ]
        },
        r".\data\sheet_wise_workbook2.xlsx": {
            "Sheet1": [
                ["Name", "Age"],
                ["Dee", 26],
                ["Alex", 40]
            ],
            "Sheet2": [
                ["City", "Country"],
                ["Berlin", "Germany"],
                ["Paris", "France"]
            ]
        }
    }

    def create_workbook(filename, sheets_data):
        workbook = Workbook()

        first_sheet_name, first_sheet_data = list(sheets_data.items())[0]
        sheet = workbook.active
        sheet.title = first_sheet_name
        for row in first_sheet_data:
            sheet.append(row)

        for sheet_name, sheet_data in list(sheets_data.items())[1:]:
            sheet = workbook.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)

        workbook.save(filename)
        workbook.close()

    for filename, sheets_data in data.items():
        create_workbook(filename, sheets_data)

    list_of_files = [r".\data\sheet_wise_workbook1.xlsx", r".\data\sheet_wise_workbook2.xlsx"]
    output_file = r".\data\merged_file_sheet_wise.xlsx"
    exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="sheet_wise", skip_bad_rows=True)
    assert_that(os.path.exists(output_file)).is_true()

    expected_data = {
                "Sheet_1" : [["Name", "Age"],["Mark", 25],["John", 30],["Dee", 26],["Alex", 40]],
                "Sheet_2" : [["City", "Country"],["New York", "USA"],["London", "UK"],["Berlin", "Germany"],["Paris","France"]]
            }

    workbook = excel.load_workbook(filename=output_file)
    sheets = workbook.sheetnames
    assert_that(sheets).is_length(2)

    for i in sheets:
        sheet = workbook[i]
        assert_that(sheet.max_row).is_equal_to(5)
        assert_that(sheet.max_column).is_equal_to(2)

        for row_index, expected_row in enumerate(expected_data[i], start=1):
            for col_index, expected_value in enumerate(expected_row, start=1):
                cell_value = sheet.cell(row=row_index, column=col_index).value
                if cell_value != expected_value:
                    assert False, f"Data mismatch ({cell_value} != {expected_value})"

    workbook.close()
    delete_the_test_excel_file(files=[output_file, r".\data\sheet_wise_workbook2.xlsx", r".\data\sheet_wise_workbook1.xlsx"])


def test_merge_excels_empty_files_list(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        exl.merge_excels(file_list=[], output_filename="output_file.xlsx")

    assert_that(str(exc_info.value)).is_equal_to("The file list is empty. Provide at least one file to merge.")


def test_merge_excels_invalid_merge_type(setup_teardown):
    with pytest.raises(ValueError) as exc_info:
        list_of_files = [r".\data\workbook1.xlsx", r".\data\workbook2.xlsx"]
        output_file = r".\data\merged_file.xlsx"
        exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="invalid_merge_type")

    assert_that(str(exc_info.value)).is_equal_to("Invalid merge type. Use 'multiple_sheets', 'single_sheet', or 'sheet_wise'.")


def test_merge_excels_multi_sheet_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        list_of_files = [INVALID_EXCEL_FILE_PATH, r".\data\workbook2.xlsx"]
        output_file = r".\data\merged_file.xlsx"
        exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="multiple_sheets")

    assert_that(str(exc_info.value)).is_equal_to(f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path.")


def test_merge_excels_single_sheet_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        list_of_files = [INVALID_EXCEL_FILE_PATH, r".\data\workbook2.xlsx"]
        output_file = r".\data\merged_file.xlsx"
        exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="single_sheet")

    assert_that(str(exc_info.value)).is_equal_to(f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path.")


def test_merge_excels_sheet_wise_file_not_found(setup_teardown):
    with pytest.raises(ExcelFileNotFoundError) as exc_info:
        list_of_files = [INVALID_EXCEL_FILE_PATH, r".\data\workbook2.xlsx"]
        output_file = r".\data\merged_file.xlsx"
        exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="sheet_wise")

    assert_that(str(exc_info.value)).is_equal_to(f"Excel file '{INVALID_EXCEL_FILE_PATH}' not found. Please give the valid file path.")


def test_merge_excels_sheet_wise_index_error(setup_teardown):
    data = {
        r".\data\sheet_wise_workbook1.xlsx": {
            "Sheet1": [
                ["Name", "Age"],
                ["Mark", 25],
                ["John", 30]
            ],
            "Sheet2": [
                ["City", "Country"],
                ["New York", "USA"],
                ["London", "UK"]
            ]
        },
        r".\data\sheet_wise_workbook2.xlsx": {
            "Sheet1": [
                ["Name", "Age"],
                ["Dee", 26],
                ["Alex", 40]
            ]
        }
    }

    def create_workbook(filename, sheets_data):
        workbook = Workbook()

        first_sheet_name, first_sheet_data = list(sheets_data.items())[0]
        sheet = workbook.active
        sheet.title = first_sheet_name
        for row in first_sheet_data:
            sheet.append(row)

        for sheet_name, sheet_data in list(sheets_data.items())[1:]:
            sheet = workbook.create_sheet(title=sheet_name)
            for row in sheet_data:
                sheet.append(row)

        workbook.save(filename)
        workbook.close()

    for filename, sheets_data in data.items():
        create_workbook(filename, sheets_data)

    with pytest.warns(UserWarning, match="Skipping"):
        list_of_files = [r".\data\sheet_wise_workbook1.xlsx", r".\data\sheet_wise_workbook2.xlsx"]
        output_file = r".\data\merged_file_sheet_wise.xlsx"
        exl.merge_excels(file_list=list_of_files, output_filename=output_file, merge_type="sheet_wise")

    delete_the_test_excel_file(files=[r".\data\sheet_wise_workbook2.xlsx", r".\data\sheet_wise_workbook1.xlsx"])
